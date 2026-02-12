import pandas as pd
from collections import Counter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ====== PATHS (popravi po potrebi) ======
# Priporočilo: nastavi absolutne poti ali uporabi Path(__file__).resolve() glede na tvoj projekt
INPUT_PATH  = Path(r"C:\Users\gl8304\Documents\Projekti\Faks\Thesis\data\context\ActivityContextGen_v12.xlsx")
OUTPUT_PATH = Path(r"C:\Users\gl8304\Documents\Projekti\Faks\Thesis\data\context\ContextGroups_act_3_4_5.xlsx")
SHEET_NAME  = "ActionLst"

# ====== Kontekstne spremenljivke (isto kot prej) ======
VAR_COLS = [
    "rec_C_T1","rec_C_T2","rec_C_T3","rec_C_P1","rec_C_P2","rec_C_P3",
    "act_C_T1","act_C_T2","act_C_T3","act_C_P1","act_C_P2","act_C_P3",
    "act_C_A1","act_C_A2","act_C_A3"
]

# ====== NOVE kontekstne grupe (iz maila) ======
GROUP_COLS = {
    "ContextG_act_3": "ContextG_act_3",
    "ContextG_act_4": "ContextG_act_4",
    "ContextG_act_5": "ContextG_act_5",
}

def clean_vals(series: pd.Series) -> list[str]:
    """Počisti vrednosti: odstrani NaN, prazno, '-1'."""
    vals = []
    for v in series.dropna().tolist():
        s = str(v).strip()
        if s == "" or s.lower() == "nan" or s == "-1":
            continue
        vals.append(s)
    return vals

def top_vals(df_sub: pd.DataFrame, col: str, k: int = 3):
    """Vrne top-k najpogostejših vrednosti v stolpcu."""
    return Counter(clean_vals(df_sub[col])).most_common(k)

def build_desc_triples(df_sub: pd.DataFrame, top_k: int = 3) -> str:
    """
    Naredi opis grupe:
      - povzetek dominantnega časa/prostora/spremljevalne aktivnosti
      - za vsak kontekstni stolpec top-k frekvence
    """
    def pick_dom(cols):
        for c in cols:
            tv = top_vals(df_sub, c, k=1)
            if tv:
                return tv[0][0]
        return None

    dom_t = pick_dom(["rec_C_T1","rec_C_T2","rec_C_T3","act_C_T1","act_C_T2","act_C_T3"])
    dom_p = pick_dom(["rec_C_P1","rec_C_P2","rec_C_P3","act_C_P1","act_C_P2","act_C_P3"])
    dom_a = pick_dom(["act_C_A1","act_C_A2","act_C_A3"])

    header = []
    if dom_t: header.append(f"dominantni čas: '{dom_t}'")
    if dom_p: header.append(f"dominantni prostor: '{dom_p}'")
    if dom_a: header.append(f"dominantna spremljevalna aktivnost: '{dom_a}'")

    lines = []
    lines.append(
        "Povzetek: " +
        ("; ".join(header) if header else "brez jasne dominante") +
        f". N={len(df_sub)} aktivnosti."
    )

    for col in VAR_COLS:
        tv = top_vals(df_sub, col, k=top_k)
        if tv:
            triples = ", ".join([f"({col}, '{v}', {n})" for v, n in tv])
            lines.append(f"{col}: [{triples}]")

    return "\n".join(lines)

# ====== LOAD ======
df = pd.read_excel(INPUT_PATH, sheet_name=SHEET_NAME)
df.columns = df.columns.astype(str).str.strip()

missing = [c for c in (VAR_COLS + list(GROUP_COLS.values())) if c not in df.columns]
if missing:
    raise ValueError(f"Manjkajo stolpci v Excelu: {missing}")

# ====== Prepare summaries ======
summaries = {}
for label, gcol in GROUP_COLS.items():
    summaries[label] = {}

    # vzemi samo ne-NaN grupe
    uniq_groups = df[gcol].dropna().unique().tolist()

    # stabilno sortiranje (da ne crkne pri mix str/int): pretvori v string za sort ključ
    uniq_groups_sorted = sorted(uniq_groups, key=lambda x: str(x))

    for grp in uniq_groups_sorted:
        df_grp = df[df[gcol] == grp]
        summaries[label][grp] = build_desc_triples(df_grp, top_k=3)

# ====== WRITE ONE-SHEET EXCEL ======
wb = Workbook()
ws = wb.active
ws.title = "ContextGroups"

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 120

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
bold = Font(bold=True)
wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")
center = Alignment(vertical="center", horizontal="center")

row = 1

def write_section(section_name: str, section_dict: dict):
    global row

    # Header vrstica
    ws.cell(row=row, column=1, value=section_name).font = bold
    ws.cell(row=row, column=1).alignment = center

    ws.cell(row=row, column=2, value="Opis").font = bold
    ws.cell(row=row, column=2).alignment = center

    for c in (1, 2):
        ws.cell(row=row, column=c).border = border
    ws.row_dimensions[row].height = 20
    row += 1

    # Vrstice za grupe
    for grp, txt in section_dict.items():
        ws.cell(row=row, column=1, value=str(grp)).alignment = center
        ws.cell(row=row, column=2, value=txt).alignment = wrap_top
        ws.cell(row=row, column=2).font = Font(size=10)

        for c in (1, 2):
            ws.cell(row=row, column=c).border = border
        ws.row_dimensions[row].height = 240
        row += 1

# piši sekcije v istem sheetu (ločeno po ContextG_act_3/4/5)
write_section("ContextG_act_3", summaries["ContextG_act_3"])
write_section("ContextG_act_4", summaries["ContextG_act_4"])
write_section("ContextG_act_5", summaries["ContextG_act_5"])

ws.freeze_panes = "A2"
wb.save(OUTPUT_PATH)

print(f"OK ✅ Ustvarjen Excel: {OUTPUT_PATH}")
